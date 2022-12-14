from __future__ import annotations
import abc
from collections.abc import Generator
from pathlib import Path

import numpy as np
import openpyxl as xl
import openpyxl.utils as xlutils


## ---------------------------------------------------------------------------------------------------------------------
## Workbook interface
## ---------------------------------------------------------------------------------------------------------------------


class WorksheetBase(abc.ABC):
    def __init__(self, workbook: Workbook, openpyxl_worksheet) -> None:
        self.workbook           = workbook
        self.openpyxl_worksheet = openpyxl_worksheet

        self.data = self.load_data()

    def load_data(self) -> np.ndarray:
        return np.array([
            [cell.value for cell in row]
            for row in self.openpyxl_worksheet.iter_rows()
        ])

    @property
    def name(self) -> str:
        return self.openpyxl_worksheet.title


class Worksheet(WorksheetBase):
    def __init__(self, workbook: Workbook, openpyxl_worksheet, *, crop_out_empty=True) -> None:
        self.crop_out_empty = crop_out_empty

        super().__init__(workbook, openpyxl_worksheet)

        if self.crop_out_empty:
            self.data = self.crop_out_empty_rows_and_columns(self.data)

    @staticmethod
    def crop_out_empty_rows_and_columns(data: np.ndarray) -> Generator[np.ndarray]:
        """Crop out empty bottom-most and right-most rows and columns, respectively.
        """
        for row_idx in range(len(data)-1, -1, -1):
            if not (data[row_idx] == None).all():
                last_nonempty_row_idx = row_idx
                break

        for col_idx in range(len(data[0])-1, -1, -1):
            if not (data[:, col_idx] == None).all():
                last_nonempty_col_idx = col_idx
                break

        return data[:last_nonempty_row_idx+1, :last_nonempty_col_idx+1]


class Workbook:
    def __init__(self, path: Path, *, read_only=True, crop_out_empty=True, worksheet_class=Worksheet):
        self.path            = path
        self.read_only       = read_only
        self.crop_out_empty  = crop_out_empty
        self.worksheet_class = worksheet_class

        assert issubclass(self.worksheet_class, WorksheetBase)

        self.worksheets = self.load_worksheets()

    def load_worksheets(self) -> dict:
        xl_wb = xl.load_workbook(self.path, data_only=True, read_only=self.read_only)
        return {
            xl_ws.title: self.worksheet_class(self, xl_ws, crop_out_empty=self.crop_out_empty)
            for xl_ws in xl_wb.worksheets
        }

    @property
    def worksheet_names(self) -> tuple[str]:
        return tuple(self.worksheets.keys())

    def __iter__(self) -> Generator[tuple[str, Worksheet]]:
        for n, ws in self.worksheets.items():
            yield n, ws

    def __getitem__(self, worksheet_name: str) -> Worksheet:
        return self.worksheets[worksheet_name]


## ---------------------------------------------------------------------------------------------------------------------
## Excel coordinate conversion functions
## ---------------------------------------------------------------------------------------------------------------------

def xlrowref(row_idx, zero_indexed=True):
    if zero_indexed:
        row_idx += 1
    return row_idx


def xlcolref(column_idx, zero_indexed=True):
    if zero_indexed:
        column_idx += 1
    return xlutils.get_column_letter(column_idx)


def xlref(row_idx, column_idx, zero_indexed=True):
    return xlcolref(column_idx, zero_indexed) + str(xlrowref(row_idx, zero_indexed))


## ---------------------------------------------------------------------------------------------------------------------
## Misc
## ---------------------------------------------------------------------------------------------------------------------


class Meta:
    def __init__(self, **kwargs):
        for key, value in kwargs.items():
            setattr(self, key, value)
